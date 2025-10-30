import logging
from io import BytesIO
from openpyxl import load_workbook
from adapters.google_storage_bucket import GoogleStorageBucket

def normalizar_celda(value):
    return None if value is None or (isinstance(value, str) and not value.strip()) else value

COLUMNAS_REQUERIDAS = [
    "zone", "project", "dataset", "table", "filter_sentence",
    "schedule", "periodicidad", "llaves", "regla_completitud", "columnas_completitud", "columna_particion", "descripcion_scan"
]

class GoogleStorageXlsxReader:
    """Lector de archivos XLSX almacenados en GCS. Busca específicamente la hoja 'Calidad Basica' y convierte sus filas en diccionarios."""

    def __init__(self, storage_bucket: GoogleStorageBucket):
        self.storage_bucket = storage_bucket

    def read_calidad_basica(self, file_name: str) -> list[dict]:
        """Lee el archivo XLSX desde GCS y devuelve una lista de diccionarios con los datos de la hoja 'Calidad Basica'"""
        file_bytes = self.storage_bucket.download_file_as_bytes(file_name)
        excel = load_workbook(filename=BytesIO(file_bytes), data_only=True)

        if "Calidad Basica" not in excel.sheetnames:
            raise ValueError("[GoogleStorageXlsxReader] El archivo no contiene la hoja 'Calidad Basica'")

        hoja = excel["Calidad Basica"]
        headers = [celda.value for celda in next(hoja.iter_rows(min_row=1, max_row=1))]

        columnas_faltantes = [col for col in COLUMNAS_REQUERIDAS if col not in headers]
        if columnas_faltantes:
            raise ValueError(f"[GoogleStorageXlsxReader] Faltan columnas obligatorias en el Excel: {columnas_faltantes}")

        data = []

        for indice, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            if all(normalizar_celda(celda) is None for celda in fila):
                break

            fila_dict = {headers[i]: normalizar_celda(fila[i]) for i in range(len(headers))}
            data.append(fila_dict)
        return data
